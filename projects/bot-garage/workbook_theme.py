"""
workbook_theme.py
=================

Portfolio excerpt, adapted. Shared openpyxl styling: one Theme plus helpers
for headers, banded tables, native Excel tables, freeze panes, drop-downs and
conditional formatting.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable, Optional, Sequence

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


# --------------------------------------------------------------------------- #
# Theme
# --------------------------------------------------------------------------- #

@dataclass(frozen=True)
class Theme:
    """Single source for the palette; edit here, not in each worksheet."""

    ink: str = "FF1A1A1A"
    paper: str = "FFFFFFFF"
    accent: str = "FFB11313"
    band: str = "FFF4F4F4"
    good_bg: str = "FFC6EFCE"
    good_fg: str = "FF006100"
    bad_bg: str = "FFFFC7CE"
    bad_fg: str = "FF9C0006"
    warn_bg: str = "FFFFEB9C"
    warn_fg: str = "FF9C6500"
    info_bg: str = "FFDDEBF7"
    info_fg: str = "FF1F4E78"


THEME = Theme()

_THIN = Side(style="thin", color="FFD0D0D0")
GRID_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=THEME.paper)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=THEME.paper)
CELL_FONT = Font(name="Calibri", size=10, color=THEME.ink)

TITLE_FILL = PatternFill("solid", fgColor=THEME.ink)
ACCENT_FILL = PatternFill("solid", fgColor=THEME.accent)
HEADER_FILL = PatternFill("solid", fgColor=THEME.ink)
BAND_FILL = PatternFill("solid", fgColor=THEME.band)

WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# --------------------------------------------------------------------------- #
# Layout helpers
# --------------------------------------------------------------------------- #

def title_block(ws: Worksheet, title: str, subtitle: str, n_cols: int,
                *, code: Optional[str] = None) -> int:
    """Render the two-row banner across n_cols; return the first free row."""
    last_col = get_column_letter(max(1, n_cols))
    ws.merge_cells(f"A1:{last_col}1")
    cell = ws["A1"]
    cell.value = f"  {title}" + (f"    [{code}]" if code else "")
    cell.font, cell.fill = TITLE_FONT, TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_col}2")
    sub = ws["A2"]
    sub.value = f"  {subtitle}"
    sub.fill = ACCENT_FILL
    sub.alignment = Alignment(horizontal="left", vertical="center")
    return 4


def header_row(ws: Worksheet, headers: Sequence[str], row: int) -> None:
    """Write the column headers at row with the header style."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL
        cell.alignment, cell.border = CENTER, GRID_BORDER
    ws.row_dimensions[row].height = 28


def set_widths(ws: Worksheet, widths: Iterable[int], start_col: int = 1) -> None:
    """Set column widths left to right from start_col."""
    for offset, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + offset)].width = width


def write_table(ws: Worksheet, rows: Sequence[Sequence], *, start_row: int,
                n_cols: int, band: bool = True,
                center_cols: Sequence[int] = ()) -> int:
    """Write body rows with borders and optional zebra striping; return the next row."""
    r = start_row
    for i, record in enumerate(rows):
        for col in range(1, n_cols + 1):
            # short rows pad with None rather than raising
            value = record[col - 1] if col - 1 < len(record) else None
            cell = ws.cell(row=r, column=col, value=value)
            cell.font = CELL_FONT
            cell.alignment = CENTER if col in center_cols else WRAP_TOP
            cell.border = GRID_BORDER
            if band and i % 2 == 1:
                cell.fill = BAND_FILL
        r += 1
    return r


def as_excel_table(ws: Worksheet, ref: str, name: str,
                   style: str = "TableStyleMedium2") -> None:
    """Wrap ref in a native Excel table so it gets a filter and live banding."""
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)


def freeze_and_filter(ws: Worksheet, header_row_idx: int, n_cols: int,
                      freeze_col: int = 1) -> None:
    """Freeze below the header and left of freeze_col, and add the auto-filter."""
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=freeze_col + 1)
    last = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_row_idx}:{last}{header_row_idx}"


# --------------------------------------------------------------------------- #
# Data validation + conditional formatting
# --------------------------------------------------------------------------- #

def add_dropdown(ws: Worksheet, col_letter: str, choices: Sequence[str],
                 first_row: int, last_row: int) -> None:
    """Restrict a column range to choices via a list drop-down."""
    # inline-list formula1, so choices can't contain commas
    dv = DataValidation(type="list", formula1='"' + ",".join(choices) + '"',
                        allow_blank=True, showDropDown=False)
    dv.errorTitle, dv.error = "Invalid value", "Pick a value from the list"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def cf_status(ws: Worksheet, cell_range: str) -> None:
    """Colour status cells by exact text match."""
    mapping = {
        "Done": (THEME.good_bg, THEME.good_fg),
        "In progress": (THEME.info_bg, THEME.info_fg),
        "Deferred": ("FFE2E2E2", "FF595959"),
    }
    for text, (bg, fg) in mapping.items():
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="equal", formula=[f'"{text}"'],
            fill=PatternFill("solid", fgColor=bg), font=Font(color=fg, bold=True)))


def cf_keyword(ws: Worksheet, cell_range: str, keyword: str = "VERIFY") -> None:
    """Warn-fill any cell whose text contains keyword."""
    # rule formula needs the top-left anchor; openpyxl shifts it down the range
    anchor = cell_range.split(":")[0]
    col = "".join(ch for ch in anchor if ch.isalpha())
    row = "".join(ch for ch in anchor if ch.isdigit())
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("{keyword}",{col}{row}))'],
        fill=PatternFill("solid", fgColor=THEME.warn_bg),
        font=Font(color=THEME.warn_fg, bold=True), stopIfTrue=False))
